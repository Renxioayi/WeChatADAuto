#coding = utf-8
from openpyxl import load_workbook
from auto import Auto
import os
import datetime

def getConditions():
    #定义conditions为list列表
    conditions = []
    try:
        #打开Excel表格
        wb = load_workbook('auto.xlsx')
        #获取当前正在显示的sheet
        sheet = wb.active

        #定义一个数组,存储对应的参数
        array = ['id','position','start_date','stop_date','is_ocpa']
        #定义行循环,循环第二行到 sheet.max+row 行
        for i in range(2,sheet.max_row + 1):
            #定义condition为字典数据类型
            condition = {}
            #使用enumerate定义列循环
            for j,arr in enumerate(array):
                #condition存入sheet.cell获取的单元格内容
                condition[arr] = sheet.cell(i,j+1).value
            #使用append将condition添加到新的对象
            conditions.append(condition)

    except FileNotFoundError:
        print("auto.xlsx文件不存在")

    return conditions

def getImageNameList():
    conditions = getConditions()
    #定义当前系统时间
    st = datetime.datetime.now().strftime('%Y-%m-%d  %H-%M-%S')
    #新建上传失败txt文本
    EN = open(st + '上传失败记录.txt','a')
    auto = Auto()

    for condition in conditions:
        #判断是否是ocpa
        if condition['is_ocpa'] == '是':
            is_ocpa = '是'
            suffix = 'ocpa'
        else:
            is_ocpa = '否'
            suffix = ''
        i = 0
        j = 0
        dirName = condition['position'] + suffix
        #传入对应的参数
        print("正在获取原始ID为" + condition['id'] +'的广告文案信息')
        x = len([name for name in os.listdir(dirName) if os.path.isfile(os.path.join(dirName, name))])
        print('共需上传' +str(x) +'个广告文案')
        for img in os.listdir(condition['position'] + suffix):
            #进行异常处理
            i += 1
            j += 1
            try:
                auto.auto_ad(condition['id'],condition['position'],condition['start_date'],condition['stop_date'],img,is_ocpa)
                print(st + '  ' + img + '  ' +"上传成功" + '  ' + '广告创建进度为: ' + str(i)+ '/' + str(x) + '  ' + "成功的为" + str(j) +'个')
            except Exception as e:
                print(st + '  ' + img + '  ' + "上传失败" + '  ' + '广告创建进度为: ' + str(i) +'/' + str(x))
                EN.write(condition['id'] + '  ' + img + '  ' + "上传失败"  + '  ' + str(i)+ '/' + str(x) + '\n')
                print(e)
                continue
        # EN.write('广告共有: ' + str(i) +"成功的为" + str(j) +'个')
        EN.close()
        # 定义循环结束页面

    input("=======================上传完成=======================")

if __name__ == '__main__':
     print("=======================欢迎使用南讯传媒微信广告服务商平台文案自动化上传脚本=======================")
     getImageNameList()












